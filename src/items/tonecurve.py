import tkinter as tk
from tkinter import Toplevel, filedialog, Canvas
from tkinter.ttk import Button
import os
import numpy as np
from PIL import Image
import cv2
import openpyxl

CHANNELS = {"R": 0, "G": 1, "B": 2}
PARAM_DIR = os.path.join(os.path.dirname(__file__), "tonecurve_params")

class TonecurveWindow:
    def __init__(self, root):
        self.root = root
        self.window = Toplevel()
        self.window.title("Tone Curve Adjustment")

        self.luts = [[] for _ in range(3)]
        self.original_image = None

        self.image_on_canvas = None

        self.hist_canvas = Canvas(self.window, width=768, height=150, bg="black")
        self.hist_canvas.pack(side=tk.TOP)

        self.r_canvas = Canvas(self.window, width=256, height=256, bg="white",
                                  #takefocus = True,
                                  )
        self.g_canvas = Canvas(self.window, width=256, height=256, bg="white",
                                  #takefocus = True,
                                  )
        self.b_canvas = Canvas(self.window, width=256, height=256, bg="white",
                                  #takefocus = True,
                                  )

        self.r_canvas.pack(side=tk.LEFT)
        self.g_canvas.pack(side=tk.LEFT)
        self.b_canvas.pack(side=tk.LEFT)

        self.tone_curve_adjusters = {
            "R": ToneCurveAdjuster(self, self.r_canvas, "R", "red"),
            "G": ToneCurveAdjuster(self, self.g_canvas, "G", "green"),
            "B": ToneCurveAdjuster(self, self.b_canvas, "B", "blue"),
        }
        for channel_name in self.tone_curve_adjusters:
            self.luts[CHANNELS[channel_name]] = self.tone_curve_adjusters[channel_name].lut.copy()

        self.reset_button = Button(self.window, text="Reset", command=self.reset_curves)
        self.reset_button.pack(side=tk.TOP)

        self.save_button = Button(self.window, text="Save", command=self.save_curves)
        self.save_button.pack(side=tk.TOP)

        self.load_button = Button(self.window, text="Load", command=self.load_curves)
        self.load_button.pack(side=tk.TOP)

        if not os.path.exists(PARAM_DIR):
            os.makedirs(PARAM_DIR)

    def set_image(self, image_pil):
        self.original_image = np.array(image_pil, dtype=np.uint8)
    
    def preprocess(self):
        self.display_image(self.original_image)
        self.draw_histogram()

    def apply_process(self):
        channels = list(cv2.split(self.original_image))
        for idx in CHANNELS.values():
            channels[idx] = cv2.LUT(channels[idx], self.luts[idx])
        adjusted_image = cv2.merge(channels)
        return Image.fromarray(adjusted_image)

    def display_image(self, image):
        self.image_on_canvas = Image.fromarray(image)
        self.root.update_image(self.image_on_canvas)

    def update_lut(self, lut, channel_idx):
        self.luts[channel_idx] = lut

    def apply_tone_curve(self, lut, channel_name):
        if self.original_image is not None:
            channels = list(cv2.split(self.original_image))
            channel_idx = CHANNELS[channel_name]
            self.update_lut(lut, channel_idx)
            for idx in CHANNELS.values():
                channels[idx] = cv2.LUT(channels[idx], self.luts[idx])
            adjusted_image = cv2.merge(channels)
            self.display_image(adjusted_image)
            self.draw_histogram(adjusted_image)

    def reset_curves(self):
        for adjuster in self.tone_curve_adjusters.values():
            adjuster.reset_curve()
        if self.original_image is not None:
            self.display_image(self.original_image)
            self.draw_histogram()

    def draw_histogram(self, image=None):
        self.hist_canvas.delete("histogram")
        if image is None:
            image = self.original_image

        # ヒストグラムを計算
        hist_r = cv2.calcHist([image], [CHANNELS["R"]], None, [256], [0, 256])
        hist_g = cv2.calcHist([image], [CHANNELS["G"]], None, [256], [0, 256])
        hist_b = cv2.calcHist([image], [CHANNELS["B"]], None, [256], [0, 256])

        max_value = max(hist_r.max(), hist_g.max(), hist_b.max())
        for i in range(256):
            r_value = int(hist_r[i][0] * 150 / max_value)
            g_value = int(hist_g[i][0] * 150 / max_value)
            b_value = int(hist_b[i][0] * 150 / max_value)

            self.hist_canvas.create_line(i*3, 150, i*3, 150-r_value, fill="red", tags="histogram")
            self.hist_canvas.create_line(i*3+1, 150, i*3+1, 150-g_value, fill="green", tags="histogram")
            self.hist_canvas.create_line(i*3+2, 150, i*3+2, 150-b_value, fill="blue", tags="histogram")

    def save_curves(self):
        file_path = filedialog.asksaveasfilename(
                                                initialdir=PARAM_DIR,
                                                initialfile="curve_points",
                                                defaultextension=".xlsx",  # デフォルトの拡張子
                                                filetypes=[("Parameter Files", "*.xlsx")],
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["channel", "x", "y"])
        for channel, adjuster in self.tone_curve_adjusters.items():
            for points in adjuster.curve_points:
                ws.append([channel, *points])
        wb.save(file_path)

    def load_curves(self):
        file_path = filedialog.askopenfilename(filetypes=[("Parameter Files", "*.xlsx")],
                                               initialdir=PARAM_DIR)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[0]

        # set curve points in each adjusters
        for adjuster in self.tone_curve_adjusters.values():
            adjuster.curve_points = []
        for param in ws.iter_rows(min_row=2, values_only=True):
            channel, x, y = param
            self.tone_curve_adjusters[channel].curve_points.append([x, y])
            
        # make tonecurves
        for adjuster in self.tone_curve_adjusters.values():
            adjuster.draw_curve()
            adjuster.update_image()

class ToneCurveAdjuster:
    def __init__(self, parent, canvas, channel_name, color):
        self.parent = parent
        self.canvas = canvas
        self.channel_name = channel_name
        self.color = color
        self.curve_points = [(0, 255), (255, 0)]
        self.selected_point = None

        self.canvas.bind("<Button-3>", self.add_point)
        self.canvas.bind("<B3-Motion>", self.move_new_point)
        self.canvas.bind("<B1-Motion>", self.move_existing_point)
        self.canvas.bind("<ButtonRelease-3>", self.update_image)
        self.canvas.bind("<ButtonRelease-1>", self.update_image)
        self.canvas.bind("<Button-2>", self.delete_selected_point)
        self.draw_curve()
        self.lut = self.generate_lut()

    def draw_curve(self):
        self.canvas.delete("curve")
        self.draw_grid()
        for i, point in enumerate(self.curve_points):
            x, y = point
            if i == self.selected_point:
                self.canvas.create_oval(x-4, y-4, x+4, y+4, outline=self.color, fill="white", width=2, tags="curve")
            else:
                self.canvas.create_oval(x-3, y-3, x+3, y+3, fill=self.color, tags="curve")
        for i in range(len(self.curve_points) - 1):
            x1, y1 = self.curve_points[i]
            x2, y2 = self.curve_points[i + 1]
            self.canvas.create_line(x1, y1, x2, y2, fill=self.color, tags="curve")

    def draw_grid(self):
        self.canvas.create_rectangle(0, 0, 256, 256, fill="darkgray", tags="curve")
        for i in range(0, 256, 10):
            color = "white" if i % 50 == 0 else "lightgray"
            self.canvas.create_line(i, 0, i, 256, fill=color, tags="curve")
            self.canvas.create_line(0, i, 256, i, fill=color, tags="curve")
        for i in range(0, 256, 50):
            self.canvas.create_text(i, 245, text=str(i), fill="white", tags="curve")
            self.canvas.create_text(15, 255-i, text=str(i), fill="white", tags="curve")

    def add_point(self, event):
        x = max(0, min(255, event.x))
        y = self.calculate_lut_value(x)
        self.curve_points.append((x, y))
        self.curve_points.sort()
        self.selected_point = self.curve_points.index((x, y))
        self.draw_curve()

    def move_new_point(self, event):
        if self.selected_point is not None and self.selected_point not in [0, len(self.curve_points) - 1]:
            min_x = self.curve_points[self.selected_point - 1][0]
            max_x = self.curve_points[self.selected_point + 1][0]
            x = max(min_x, min(max_x, event.x))
            if event.state & 0x4: # Ctrlで分岐
                y = 255 - self.lut[x]
            else:
                y = max(0, min(255, event.y))
            self.curve_points[self.selected_point] = (x, y)
            self.draw_curve()

    def move_existing_point(self, event):
        if len(self.curve_points) > 2:
            self.selected_point = self.get_nearest_point(event.x, event.y)

            min_x = self.curve_points[self.selected_point - 1][0]
            max_x = self.curve_points[self.selected_point + 1][0]
            x = max(min_x, min(max_x, event.x))

            y = max(0, min(255, event.y))
            self.curve_points[self.selected_point] = (x, y)
            self.draw_curve()

    def get_nearest_point(self, x, y):
        min_dist = float('inf')
        nearest_point = None
        curve_points_without_edge = self.curve_points[1:-1]
        for i, (px, py) in enumerate(curve_points_without_edge):
            dist = (px - x) ** 2 + (py - y) ** 2
            if dist < min_dist:
                min_dist = dist
                nearest_point = i + 1 # (0, 0)を除くため
        return nearest_point

    def update_image(self, event=None):
        self.lut = self.generate_lut()
        self.parent.apply_tone_curve(self.lut, self.channel_name)

    def generate_lut(self):
        lut = np.zeros(256, dtype=np.uint8)
        for i in range(256):
            lut[i] = self.calculate_lut_value(i)
        return lut

    def calculate_lut_value(self, x):
        for i in range(len(self.curve_points) - 1):
            x1, y1 = self.curve_points[i]
            x2, y2 = self.curve_points[i + 1]
            y1, y2 = 255 - y1, 255 - y2 # x=yに対して対称
            if x1 <= x <= x2:
                t = (x - x1) / (x2 - x1)
                return int(y1 * (1 - t) + y2 * t)
        return 0

    def reset_curve(self):
        self.curve_points = [(0, 255), (255, 0)]
        self.draw_curve()
        self.update_image()

    def delete_selected_point(self, event):
        if self.selected_point is not None:
            self.curve_points.pop(self.selected_point)
            self.draw_curve()
            self.update_image()
            self.selected_point = None

if __name__ == "__main__":
    # Main application
    root = tk.Tk()
    app = TonecurveWindow(None, root)  # Test without an image.
    root.mainloop()
